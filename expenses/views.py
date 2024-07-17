from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.paginator import Paginator
# from django.core.cache import cache
from django.template.loader import render_to_string
from django.utils.translation import gettext as _
from django.utils import translation
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum
from preferences.models import UserPreferences
from .models import Category, Expense
from .forms import UploadFileForm
from weasyprint import HTML
import datetime, json, os, csv, xlrd, tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from io import BytesIO

# Create your views here.

def index(request):
    if request.user.is_authenticated:
        expenses = Expense.objects.filter(user=request.user)
        paginator = Paginator(expenses, 5)

        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        preferences = UserPreferences.objects.filter(user=request.user)[0] # Get the currency from the UserPreferences model.
        preferences = preferences.currency[:3]
        language = UserPreferences.objects.get(user=request.user).language

        return render(request, 'expenses/index.html', {'page_obj': page_obj, 'preferences': preferences, 'language': language}) # Send the prefered currency from preferences also

    return render(request, 'expenses/index.html')

def add_expense(request):
    
    # Grabs each on of the categories in the Category model.
    categories = Category.objects.all()

    if request.method == 'GET':

        return render(request, 'expenses/add_expense.html', {'categories': categories})
    
    elif request.method == 'POST':

        # WHEN A USER IS AUTHENTICATED.
        if request.user.is_authenticated:
            # If no description is provided we just do a default.
            descriptionValue = _("No description provided.")

            # In case the user uses spaces in the description we still don't count them as a description.
            if request.POST['description'].strip():
                descriptionValue = request.POST['description']

            ## GET ALL THE LANGUAGES FOR THE CATEGORY.
            categoriesLang = Category.objects.get(name=request.POST['category'])

            newExpensesList = Expense.objects.create(
                user = request.user,
                name = request.POST['expenseName'],
                date = request.POST['datePicked'],
                description = descriptionValue,
                amount = request.POST['amount'],
                category = categoriesLang.name_en,
                category_en = categoriesLang.name_en,
                category_ja = categoriesLang.name_ja,
                category_es = categoriesLang.name_es,
            )
            newExpensesList.save()

            messages.success(request, _("Expense Added to your list!"))
            return redirect('expenses')
        
        # When the user is not authenticated.
        else:
            messages.success(request, _("Expense Added to your list!"))
            return redirect('expenses')

def delete_expense(request,pk):
    """
    Given an ID it deletes that item from the database.
    """
    if request.method == 'POST':
        expense = Expense.objects.get(id=pk)
        expense.delete()
        messages.success(request, _("Expense deleted successfully!"))
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

def edit_expense(request, pk):
    # Grabs each on of the categories in the Category model.
    categories = Category.objects.all()
    
    if request.user.is_authenticated:
        expense = Expense.objects.get(id=pk)

        if request.user == expense.user:
            if request.method == 'GET':

                return render(request, 'expenses/edit_expense.html', {'categories': categories, 'expenses': expense})
            
            elif request.method == 'POST':
                
                # If no description is provided we just do a default.
                descriptionValue = _("No description provided.")

                # In case the user uses spaces in the description we still don't count them as a description.
                if request.POST['description'].strip():
                    descriptionValue = request.POST['description']

                # WHEN A USER IS AUTHENTICATED.
                if request.user.is_authenticated:

                    editedExpense = Expense.objects.get(id=pk)

                    ## GET ALL THE LANGUAGES FOR THE CATEGORY.
                    categoriesLang = Category.objects.get(name=request.POST['category'])
                        
                    editedExpense.name = request.POST['expenseName']
                    editedExpense.date = request.POST['datePicked']
                    editedExpense.description = descriptionValue
                    editedExpense.amount = request.POST['amount']
                    editedExpense.category = categoriesLang.name_en,
                    editedExpense.category_en = categoriesLang.name_en
                    editedExpense.category_ja = categoriesLang.name_ja
                    editedExpense.category_es = categoriesLang.name_es

                    editedExpense.save()
                    
                messages.success(request, _("Expense edited successfully!"))
                return redirect('expenses')
        else:
            messages.error(request, _("Naughty Naughty. You don't have permissions to see that."))
            return redirect('expenses')

    else:
        ## DO SOMETHING WHEN THE USER IS NOT AUTHENTICATED
        if request.method == 'GET':
            return render(request, 'expenses/edit_expense.html', {'categories': categories, 'expenses': pk})
        else:
            messages.success(request, _("Expense edited successfully!"))
            return redirect('expenses')
 
def search_expense(request):

    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        search_str = data.get('searchText', '')

        # In case we use redis cache.
        # cache_key = f'search_expense_{request.user.id}_{search_str}'
        # cached_result = cache.get(cache_key)

        # if cached_result is not None:
        #     return JsonResponse(cached_result, safe=False)

        expenses = Expense.objects.filter(
            amount__startswith=search_str, user=request.user) | Expense.objects.filter(
                date__startswith=search_str, user=request.user) | Expense.objects.filter(
                    description__icontains=search_str, user=request.user) | Expense.objects.filter(
                        name__icontains=search_str, user=request.user) | Expense.objects.filter(
                            category__istartswith=search_str, user=request.user)
        
        preferences = UserPreferences.objects.filter(user=request.user)[0] # Get the currency from the UserPreferences model.
        
        data = list(expenses.values())
        if data:
            data[0]['currency'] = preferences.currency[:3]
        
        # cache.set(cache_key, data, timeout=300)  # Cache the result for 5 minutes (300 seconds)
        return JsonResponse(data, safe=False)
        
    return JsonResponse({'error': _('Invalid request method')}, status=400)

def get_category(expense):
    return expense.category

def get_category_amount(category, expenses):
    amount = 0
    byCategory = expenses.filter(category=category)

    for item in byCategory:
        amount += item.amount

    return amount

def get_categories(request):
    if request.method == "GET":
        categories = Category.objects.all().values('id', 'name', 'name_en', 'name_es', 'name_ja')
        categories_list = list(categories)  # Convert queryset to list
        return JsonResponse(categories_list, safe=False)

def expenses_data(request):
    
    if request.method == 'GET':
        today = datetime.date.today()
        # lastMonth = today.datetime.timedelta(days = 30)
        # last6months = today.datetime.timedelta(days = 30*6)
        # lastyear = today.datetime.timedelta(days = 30*12)
        if request.user.is_authenticated:
            expenses = Expense.objects.filter(user=request.user)

            if expenses:

                finalRepresentation = {}

                categoryList = list(set(map(get_category, expenses)))

                for x in expenses:
                    for y in categoryList:
                        finalRepresentation[y] = get_category_amount(y, expenses)

                return JsonResponse({'expense_data': finalRepresentation}, safe=False)
            else:
                return JsonResponse({'error': _('No expenses to show.')}, status=404)
        
        else:
            ## DO SOMETHING IF THE USER IS NOT AUTHENTICATED
            return JsonResponse({'error': _('No expenses to show.')}, status=404)
        
        return JsonResponse({'error': _('No expenses to show.')}, status=404)

def expenses_summary(request):
    if request.method == 'GET':
        if request.user.is_authenticated:
            expenses = Expense.objects.filter(user=request.user)
            
            if expenses:
                return render(request, 'expenses/expenses_summary.html', {'expenses': True})
            
            else:
                return render(request, 'expenses/expenses_summary.html')
        
        else:
            ## DO SOMETHING IF THE USER IS NOT AUTHENTICATED
            return render(request, 'expenses/expenses_summary.html')
    
    return render(request, 'expenses/expenses_summary.html')

def export_data(request, non_user_data = None):
    if request.user.is_authenticated:
        expenses = Expense.objects.filter(user=request.user)
        fileType = request.POST.get('filetype')
        language = translation.get_language()
        timeDate = str(datetime.datetime.now().date())

        match fileType:
            case 'csv':
                response = HttpResponse(content_type='text/csv')
                if language == 'ja':
                    response['Content-Disposition'] = f'attachment; filename = Expenses-{timeDate}.csv'
                else:
                    response['Content-Disposition'] = _('attachment; filename = Expenses-{}.csv').format(timeDate)

                writer = csv.writer(response)
                writer.writerow([_('Name'), _('Category'), _('Amount'), _('Description'), _('Date')])

                for expense in expenses:
                    writer.writerow([expense.name, expense.category, expense.amount, expense.description, expense.date])

                return response
            
            case 'xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                if language == 'ja':
                    response['Content-Disposition'] = f'attachment; filename=Expenses-{timeDate}.xlsx'
                else:
                    response['Content-Disposition'] = _('attachment; filename=Expenses-{}.xlsx').format(timeDate)

                wb = Workbook()
                ws = wb.active
                ws.title = 'Expenses'

                # Define the header
                columns = [_('Name'), _('Category'), _('Amount'), _('Description'), _('Date')]

                # Apply the header
                header_font = Font(bold=True)
                for col_num, column_title in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=column_title)
                    cell.font = header_font

                # Write data rows
                rows = expenses.values_list('name', 'category', 'amount', 'description', 'date')

                for row_num, row in enumerate(rows, 2):
                    for col_num, cell_value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num, value=str(cell_value))

                wb.save(response)
                return response

            
            case 'pdf':
                response = HttpResponse(content_type='text/pdf')
                if language == 'ja':
                    response['Content-Disposition'] = f'attachment; filename = Expenses-{timeDate}.pdf'
                else:
                    response['Content-Disposition'] = _('attachment; filename = Expenses-{}.pdf').format(timeDate)
                response['Content-Transfer-Encoding'] = 'binary'

                html_string = render_to_string('expenses/pdf-output.html', {'expenses': expenses, 'total': expenses.aggregate(Sum('amount'))['amount__sum']})
                html = HTML(string=html_string)

                result = html.write_pdf()

                with tempfile.NamedTemporaryFile(delete=True) as product:
                    product.write(result)
                    product.flush()

                    product= open(product.name, 'rb')
                    response.write(product.read())

                return response

        return HttpResponse(_("Export format not supported"))
    
    else:
        ## DO SOMETHING IF THE USER IS NOT AUTHENTICATED
        data = json.loads(request.body)
        expenses = data.get('expenses', [])
        fileType = data.get('format')
        # Get Language.
        language = translation.get_language()

        match fileType:
            case 'csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = _('attachment; filename = Expenses-{}.csv').format(str(datetime.datetime.now().date()))

                writer = csv.writer(response)
                writer.writerow([_('Name'), _('Category'), _('Amount'), _('Description'), _('Date')])

                for expense in expenses:
                    if language == 'es':
                        writer.writerow([expense['name'], expense['category_es'], expense['amount'], expense['description'], expense['date']])
                    elif language == 'ja':
                        writer.writerow([expense['name'], expense['category_ja'], expense['amount'], expense['description'], expense['date']])
                    else:
                        writer.writerow([expense['name'], expense['category'], expense['amount'], expense['description'], expense['date']])

                return response
            
            case 'xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = _('attachment; filename=Expenses-{}.xlsx').format(str(datetime.datetime.now().date()))

                wb = Workbook()
                ws = wb.active
                ws.title = 'Expenses'

                # Define the header
                columns = [_('Name'), _('Category'), _('Amount'), _('Description'), _('Date')]

                # Apply the header
                header_font = Font(bold=True)
                for col_num, column_title in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=column_title)
                    cell.font = header_font

                # Write data rows
                for row_num, expense in enumerate(expenses, 2):
                    ws.cell(row=row_num, column=1, value=expense['name'])
                    if language == 'es':
                        ws.cell(row=row_num, column=2, value=expense['category_es'])
                    elif language == 'ja':
                        ws.cell(row=row_num, column=2, value=expense['category_ja'])
                    else:
                        ws.cell(row=row_num, column=2, value=expense['category'])
                    ws.cell(row=row_num, column=3, value=expense['amount'])
                    ws.cell(row=row_num, column=4, value=expense['description'])
                    ws.cell(row=row_num, column=5, value=expense['date'])

                wb.save(response)
                return response

            case 'pdf':
                response = HttpResponse(content_type='text/pdf')
                response['Content-Disposition'] = _('attachment; filename = Expenses-{}.pdf').format(str(datetime.datetime.now().date()))
                response['Content-Transfer-Encoding'] = 'binary'

                expenses = data.get('expenses')

                totalAmount = 0.00
                
                for expense in expenses:
                    totalAmount += float(expense['amount'])
                    if language == 'es':
                        expense['category'] = expense['category_es']
                    elif language == 'ja':
                        expense['category'] = expense['category_ja']
                    else:
                        expense['category'] = expense['category']

                html_string = render_to_string('expenses/pdf-output.html', {'expenses': expenses, 'total': totalAmount})
                html = HTML(string=html_string)

                result = html.write_pdf()

                with tempfile.NamedTemporaryFile(delete=True) as product:
                    product.write(result)
                    product.flush()

                    product= open(product.name, 'rb')
                    response.write(product.read())

                return response

        return HttpResponse("Export format not supported")

    return HttpResponse("Export format not supported")

def import_data(request):
    if request.method == 'POST':
        if request.user.is_authenticated:
            file = request.FILES.get('file')
            delete_previous = request.POST.get('delete_previous') == 'true'
            # print(delete_previous)

            if not file:
                return JsonResponse({'error': _('No file provided')}, status=400)

            file_type = file.name.split('.')[-1].lower()

            try:
                if file_type == 'csv':
                    reader = csv.DictReader(file.read().decode('utf-8').splitlines())
                    df = list(reader)
                elif file_type == 'xls':
                    file_content = file.read()
                    book = xlrd.open_workbook(file_contents=file_content)
                    sheet = book.sheet_by_index(0)
                    headers = sheet.row_values(0)
                    df = [
                        {headers[i]: sheet.cell_value(row, i) for i in range(sheet.ncols)}
                        for row in range(1, sheet.nrows)
                    ]
                elif file_type == 'xlsx':
                    file_content = file.read()
                    wb = load_workbook(BytesIO(file_content))
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    df = [
                        {headers[i]: cell.value for i, cell in enumerate(row)}
                        for row in ws.iter_rows(min_row=2, values_only=False)
                    ]
                else:
                    messages.error(request, _('Unsupported file format'))
                    return JsonResponse({'error': _('Unsupported file format')})
                
                # Check that the file has all the needed columns. From each language.
                if df[0].get('Name') is not None:
                    required_columns = ['Name', 'Category', 'Amount', 'Description', 'Date']
                    lang = 'en'
                
                elif df[0].get('Nombre') is not None:
                    required_columns = ['Nombre', 'Categoría', 'Monto', 'Descripción', 'Fecha']
                    lang = 'es'
                
                elif df[0].get('名前') is not None:
                    required_columns = ['名前', 'カテゴリー', '金額', '説明', '日付']
                    lang = 'ja'
                
                else:
                    messages.error(request, _('Wrong amount of columns or names.'))
                    return JsonResponse({'error': _('Wrong amount of columns or names.')}, status=400)

                if not all(column in df[0] for column in required_columns):
                    messages.error(request, _('Wrong amount of columns or names.'))
                    return JsonResponse({'error': _('Wrong amount of columns or names.')}, status=400)
                
                # Clear existing records if the user requested it.
                if delete_previous:
                    Expense.objects.filter(user=request.user).delete()

                # We get the language of the page and change it temporarily to the one of the file if needed to be able to import the data.
                currentLang = translation.get_language()
                
                if currentLang != lang:
                    translation.activate(lang)

                # print(df)

                # Process each row in the dataframe.
                for row in df:
                    ## GET ALL THE LANGUAGES FOR THE CATEGORY.
                    categoryLang = Category.objects.get(name=row[required_columns[1]])
                    Expense.objects.create(
                        user=request.user,
                        name=row[required_columns[0]],
                        category=categoryLang.name_en,
                        category_en=categoryLang.name_en,
                        category_es=categoryLang.name_es,
                        category_ja=categoryLang.name_ja,
                        amount=row[required_columns[2]],
                        description=row[required_columns[3]],
                        date=row[required_columns[4]],
                    )
                
                # If the page language was changed we set it back to what it was.
                if currentLang != lang:
                    translation.activate(currentLang)

                messages.success(request, _('Expenses imported successfully'))
                return JsonResponse({'success': True})

            except Exception as e:
                messages.error(request, _('Error processing file: {}').format(e))
                return JsonResponse({'error': _('Error processing file: {}').format(e)}, status=400)
        
        else:
            form = UploadFileForm()
    else:
        form = UploadFileForm()