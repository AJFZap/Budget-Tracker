from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext as _
from django.utils import translation
from preferences.models import UserPreferences
from expenses.models import Category, Expense
from income.models import Source, Income
from .forms import UploadFileForm
from itertools import chain
from weasyprint import HTML
from operator import attrgetter
import datetime, json, os, csv, xlrd, tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from io import BytesIO

# Create your views here.

def index(request):
    """
    Returns the dashboard page with the needed data from the database and the preferences if the user is authenticated.
    """
    if request.user.is_authenticated:
        # Get the prefered currency from the user and if it's new and doesn't have any we create the database with the defaults.
        if UserPreferences.objects.filter(user=request.user).exists():
            preferences = UserPreferences.objects.filter(user=request.user)[0] # Get the currency from the UserPreferences model.
            
        else:
            UserPreferences.objects.create(user=request.user)
            preferences = UserPreferences.objects.filter(user=request.user)[0] # Get the currency from the UserPreferences model.

        preferences = preferences.currency[:3]

        # Get the expenses and income from the user.
        expenses = Expense.objects.filter(user=request.user)
        income = Income.objects.filter(user=request.user)

        expenseAmount = get_expenses_amount(expenses)
        incomeAmount = get_income_amount(income)
        balance = float(incomeAmount) - expenseAmount

        # Fetch the latest 5 entries from each model
        latest_expenses = Expense.objects.filter(user=request.user).order_by('-date')[:5]
        latest_incomes = Income.objects.filter(user=request.user).order_by('-date')[:5]

        # Combine and sort by date
        latest_entries = sorted(
        chain(latest_expenses, latest_incomes),
        key=attrgetter('date'),
        reverse=True
        )

        # Get the 5 most recent combined entries
        latest_entries = latest_entries[:5]

        # Prepare context.
        context = {
            'Balance': format(balance, ".2f"),
            'Expenses': format(expenseAmount, ".2f"),
            'Income': incomeAmount,
            'latest_entries': latest_entries,
            'language': UserPreferences.objects.get(user=request.user).language,
        }

        return render(request, 'dashboard/dashboard.html', {'balances': context, 'preferences': preferences})

    context = {'language': translation.get_language()}
    return render(request, 'dashboard/dashboard.html', {'balances': context})

def get_expenses_amount(expense):
    """
    Returns the total amount of the expenses
    """
    amount = 0
    
    for item in expense:
        amount += item.amount
    
    return amount

def get_income_amount(income):
    """
    Returns the total amount of the income.
    """
    amount = 0
    
    for item in income:
        amount += item.amount
    
    return amount

def export_data(request):
    """
    Export the data from both databases together in a csv, excel or pdf file.
    For both Authenticated and Non-Authenticated users.
    """
    if request.user.is_authenticated:
        expenses = Expense.objects.filter(user=request.user).exists()
        income = Income.objects.filter(user=request.user).exists()
        fileType = request.POST.get('filetype')
        
        # Checks if the user has both tables and if he doesn't it acts accordingly.
        if expenses and income:

            # Fetch all the entries from each model.
            expenses = Expense.objects.filter(user=request.user).order_by('-date')
            incomes = Income.objects.filter(user=request.user).order_by('-date')

            # Combine and sort by date.
            entries = sorted(
            chain(expenses, incomes),
            key=attrgetter('date'),
            reverse=True
            )
            total = float(get_income_amount(incomes)) - get_expenses_amount(expenses)
        
        elif expenses:
            entries = Expense.objects.filter(user=request.user)
            total = get_expenses_amount(expenses)

        elif income:
            entries = Income.objects.filter(user=request.user)
            total = float(get_income_amount(incomes))

        match fileType:
            case 'csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename = BudgetTracker-{str(datetime.datetime.now().date())}.csv'

                writer = csv.writer(response)
                writer.writerow([_('Name'), _('Source/Category'), _('Date'), _('Description'), _('Type'), _('Amount')])

                for entry in entries:
                    if entry.entry_type == _('Expense'):
                        writer.writerow([entry.name, entry.category, entry.date, entry.description, entry.entry_type, entry.amount])
                    else:
                        writer.writerow([entry.name, entry.source, entry.date, entry.description, entry.entry_type, entry.amount])
                
                return response
            
            case 'xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename=BudgetTracker-{str(datetime.datetime.now().date())}.xlsx'

                wb = Workbook()
                ws = wb.active
                ws.title = 'BudgetTracker'

                columns = [_('Name'), _('Source/Category'), _('Date'), _('Description'), _('Type'), _('Amount')]

                # Writing the header
                for col_num, column_title in enumerate(columns, 1):
                    ws.cell(row=1, column=col_num, value=column_title)

                # Holds all the needed data entries for exportation.
                data = []
                
                for entry in entries:
                    if entry.entry_type == _('Expense'):
                        data.append([entry.name, entry.category, entry.date, entry.description, entry.entry_type, entry.amount])
                    else:
                        data.append([entry.name, entry.source, entry.date, entry.description, entry.entry_type, entry.amount])
                
                # Writing the data
                for row_num, row_data in enumerate(data, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=cell_value)

                wb.save(response)

                return response

            
            case 'pdf':
                response = HttpResponse(content_type='text/pdf')
                response['Content-Disposition'] = f'attachment; filename = BudgetTracker-{str(datetime.datetime.now().date())}.pdf'
                response['Content-Transfer-Encoding'] = 'binary'

                html_string = render_to_string('dashboard/pdf-output.html', {'entries': entries, 'total': total})
                html = HTML(string=html_string)

                result = html.write_pdf()

                with tempfile.NamedTemporaryFile(delete=True) as product:
                    product.write(result)
                    product.flush()

                    product= open(product.name, 'rb')
                    response.write(product.read())

                return response

        return HttpResponse(_("Export format not supported"))
    
    # Un-Authenticated Export.
    else:
        data = json.loads(request.body)
        entries = data.get('allEntries', [])
        fileType = data.get('format')
        # Get Language.
        language = translation.get_language()

        match fileType:
            case 'csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename = BudgetTracker-{str(datetime.datetime.now().date())}.csv'

                writer = csv.writer(response)
                writer.writerow([_('Name'), _('Source/Category'), _('Date'), _('Description'), _('Type'), _('Amount')])

                for entry in entries:
                    if entry['db_type'] == 'Expense' or entry['db_type'] == 'Gasto' or entry['db_type'] == '経費':
                        if language == 'es':
                            writer.writerow([entry['name'], entry['category_es'], entry['date'], entry['description'], entry['db_type'], entry['amount']])
                        elif language == 'ja':
                            writer.writerow([entry['name'], entry['category_ja'], entry['date'], entry['description'], entry['db_type'], entry['amount']])
                        else:
                            writer.writerow([entry['name'], entry['category_en'], entry['date'], entry['description'], entry['db_type'], entry['amount']])
                    else:
                        if language == 'es':
                            writer.writerow([entry['name'], entry['source_es'], entry['date'], entry['description'], entry['db_type'], entry['amount']])
                        elif language == 'ja':
                            writer.writerow([entry['name'], entry['source_ja'], entry['date'], entry['description'], entry['db_type'], entry['amount']])
                        else:
                            writer.writerow([entry['name'], entry['source_en'], entry['date'], entry['description'], entry['db_type'], entry['amount']])

                return response
            
            case 'xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename=BudgetTracker-{str(datetime.datetime.now().date())}.xlsx'

                wb = Workbook()
                ws = wb.active
                ws.title = 'BudgetTracker'

                # Define the header
                columns = [_('Name'), _('Source/Category'), _('Date'), _('Description'), _('Type'), _('Amount')]

                # Apply the header
                header_font = Font(bold=True)
                for col_num, column_title in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=column_title)
                    cell.font = header_font

                # Write data rows
                for row_num, entry in enumerate(entries, 2):
                    ws.cell(row=row_num, column=1, value=entry['name'])
                    
                    if entry['db_type'] in ['Expense', 'Gasto', '経費']:
                        if language == 'es':
                            ws.cell(row=row_num, column=2, value=entry['category_es'])
                        elif language == 'ja':
                            ws.cell(row=row_num, column=2, value=entry['category_ja'])
                        else:
                            ws.cell(row=row_num, column=2, value=entry['category_en'])
                    else:
                        if language == 'es':
                            ws.cell(row=row_num, column=2, value=entry['source_es'])
                        elif language == 'ja':
                            ws.cell(row=row_num, column=2, value=entry['source_ja'])
                        else:
                            ws.cell(row=row_num, column=2, value=entry['source_en'])
                            
                    ws.cell(row=row_num, column=3, value=entry['date'])
                    ws.cell(row=row_num, column=4, value=entry['description'])
                    ws.cell(row=row_num, column=5, value=entry['db_type'])
                    ws.cell(row=row_num, column=6, value=entry['amount'])

                wb.save(response)
                return response

            
            case 'pdf':
                response = HttpResponse(content_type='text/pdf')
                response['Content-Disposition'] = f'attachment; filename = BudgetTracker-{str(datetime.datetime.now().date())}.pdf'
                response['Content-Transfer-Encoding'] = 'binary'

                entries = data.get('allEntries')

                totalAmount = 0.00
                
                for entry in entries:
                    if entry['db_type'] == 'Expense' or entry['db_type'] == 'Gasto' or entry['db_type'] == '経費':
                        totalAmount -= float(entry['amount'])
                        if language == 'es':
                            entry['source'] = entry['category_es']
                        elif language == 'ja':
                            entry['source'] = entry['category_ja']
                        else:
                            entry['source'] = entry['category_en']
                    else:
                        totalAmount += float(entry['amount'])
                        if language == 'es':
                            entry['source'] = entry['source_es']
                        elif language == 'ja':
                            entry['source'] = entry['source_ja']
                        else:
                            entry['source'] = entry['source_en']

                html_string = render_to_string('dashboard/pdf-output.html', {'entries': entries, 'total': totalAmount})
                html = HTML(string=html_string)

                result = html.write_pdf()

                with tempfile.NamedTemporaryFile(delete=True) as product:
                    product.write(result)
                    product.flush()

                    product= open(product.name, 'rb')
                    response.write(product.read())

                return response
        
    return HttpResponse(_("Export format not supported"))

def import_data(request):
    """
    Imports the data from csv and excel files into both databases.
    Only for Authenticated users.
    """
    if request.method == 'POST':
        if request.user.is_authenticated:
            file = request.FILES.get('file')
            delete_previous = request.POST.get('delete_previous') == 'true'

            if not file:
                return JsonResponse({'error': _('No file provided')}, status=400)

            file_type = file.name.split('.')[-1].lower()

            try:
                if file_type == 'csv':
                    reader = csv.DictReader(file.read().decode('utf-8').splitlines())
                    df = list(reader)
                elif file_type == 'xls':
                    file_content = file.read()
                    book = xlrd.open_workbook(file_contents=file_content)
                    sheet = book.sheet_by_index(0)
                    headers = sheet.row_values(0)
                    df = [
                        {headers[i]: sheet.cell_value(row, i) for i in range(sheet.ncols)}
                        for row in range(1, sheet.nrows)
                    ]
                elif file_type == 'xlsx':
                    file_content = file.read()
                    wb = load_workbook(BytesIO(file_content))
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    df = [
                        {headers[i]: cell.value for i, cell in enumerate(row)}
                        for row in ws.iter_rows(min_row=2, values_only=False)
                    ]
                else:
                    messages.error(request, _('Unsupported file format'))
                    return JsonResponse({'error': _('Unsupported file format')})
                
                print(df)

                # Check that the file has all the needed columns. From each language.
                if df[0].get('Name') is not None:
                    required_columns = ['Name', 'Source/Category', 'Date', 'Description', 'Type', 'Amount']
                    lang = 'en'
                elif df[0].get('Nombre') is not None:
                    required_columns = ['Nombre', 'Fuente/Categoria', 'Fecha', 'Descripción', 'Tipo', 'Monto']
                    lang = 'es'
                elif df[0].get('名前') is not None:
                    required_columns = ['名前', 'ソース/カテゴリ', '日付', '説明', 'タイプ', '金額']
                    lang = 'ja'
                else:
                    messages.error(request, _('Wrong amount of columns or names.'))
                    return JsonResponse({'error': _('Wrong amount of columns or names.')}, status=400)

                if not all(column in df[0] for column in required_columns):
                    messages.error(request, _('Wrong amount of columns or names.'))
                    return JsonResponse({'error': _('Wrong amount of columns or names.')}, status=400)

                # Clear existing records if the user requested it.
                if delete_previous:
                    Expense.objects.filter(user=request.user).delete()
                    Income.objects.filter(user=request.user).delete()

                # We get the language of the page and change it temporarily to the one of the file if needed to be able to import the data.
                currentLang = translation.get_language()
                if currentLang != lang:
                    translation.activate(lang)

                for row in df:
                    if row[_('Type')] == _('Expense'):
                        ## GET ALL THE LANGUAGES FOR THE CATEGORY.
                        categoryLang = Category.objects.get(name=row[required_columns[1]])
                        Expense.objects.create(
                            user=request.user,
                            name=row[_('Name')],
                            category=categoryLang.name_en,
                            category_en=categoryLang.name_en,
                            category_es=categoryLang.name_es,
                            category_ja=categoryLang.name_ja,
                            amount=row[_('Amount')],
                            description=row[_('Description')],
                            date=row[_('Date')],
                        )
                    else:
                        ## GET ALL THE LANGUAGES FOR THE SOURCE.
                        sourceLang = Source.objects.get(name=row[required_columns[1]])
                        Income.objects.create(
                            user=request.user,
                            name=row[_('Name')],
                            source=sourceLang.name_en,
                            source_en=sourceLang.name_en,
                            source_es=sourceLang.name_es,
                            source_ja=sourceLang.name_ja,
                            amount=row[_('Amount')],
                            description=row[_('Description')],
                            date=row[_('Date')],
                        )

                # If the page language was changed we set it back to what it was.
                if currentLang != lang:
                    translation.activate(currentLang)

                messages.success(request, _('Data imported successfully'))
                return JsonResponse({'success': True})

            except Exception as e:
                messages.error(request, _('Error processing file: {}').format(e))
                return JsonResponse({'error': _('Error processing file: {}').format(e)}, status=400)
    else:
        form = UploadFileForm()
        return render(request, 'upload_form.html', {'form': form})

def error_404_view(request, exception):
    """
    Returns our custom 404 page.
    """
    return render(request, '404.html')