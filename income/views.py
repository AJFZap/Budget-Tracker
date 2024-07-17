from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils.translation import gettext as _
from django.utils import translation
# from django.core.cache import cache
from django.template.loader import render_to_string
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum
from preferences.models import UserPreferences
from .models import Source, Income
from .forms import UploadFileForm
from weasyprint import HTML
import datetime, json, os, csv, xlrd, tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from io import BytesIO

def income(request):
    if request.user.is_authenticated:
        income = Income.objects.filter(user=request.user)
        paginator = Paginator(income, 5)

        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        preferences = UserPreferences.objects.filter(user=request.user)[0] # Get the currency from the UserPreferences model.
        preferences = preferences.currency[:3]
        language = UserPreferences.objects.get(user=request.user).language

        return render(request, 'income/index.html', {'page_obj': page_obj, 'preferences': preferences, 'language': language}) # Send the prefered currency from preferences also

    language = translation.get_language()
    return render(request, 'income/index.html', {'language': language})

def add_income(request):
    
    # Grabs each on of the categories in the Category model.
    sources = Source.objects.all()

    if request.method == 'GET':

        return render(request, 'income/add_income.html', {'sources': sources})
    
    elif request.method == 'POST':
        
        # If no description is provided we just do a default.
        descriptionValue = _("No description provided.")

        # In case the user uses spaces in the description we still don't count them as a description.
        if request.POST['description'].strip():
            descriptionValue = request.POST['description']

        # WHEN A USER IS AUTHENTICATED.
        if request.user.is_authenticated:
            
            ## GET ALL THE LANGUAGES FOR THE SOURCE.
            sourcesLang = Source.objects.get(name=request.POST['source'])

            newIncomeList = Income.objects.create(
                user = request.user,
                name = request.POST['incomeName'],
                date = request.POST['datePicked'],
                description = descriptionValue,
                amount = request.POST['amount'],
                source = request.POST['source'],
                source_en = sourcesLang.name_en,
                source_ja = sourcesLang.name_ja,
                source_es = sourcesLang.name_es,
            )
            newIncomeList.save()

            messages.success(request, _("Income Added to your list!"))
            return redirect('income')
        
        # When the user is not authenticated.
        else:
            messages.success(request, _("Income Added to your list!"))
            return redirect('income')

def delete_income(request,pk):
    """
    Given an ID it deletes that item from the database.
    """
    if request.method == 'POST':
        income = Income.objects.get(id=pk)
        income.delete()
        messages.success(request, _("Income deleted successfully!"))
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

def edit_income(request, pk):
    # Grabs each on of the sources in the Source model.
    sources = Source.objects.all()
    
    if request.user.is_authenticated:
        income = Income.objects.get(id=pk)

        if request.user == income.user:
            if request.method == 'GET':

                return render(request, 'income/edit_income.html', {'sources': sources, 'income': income})
            
            elif request.method == 'POST':
                
                # If no description is provided we just do a default.
                descriptionValue = _("No description provided.")

                # In case the user uses spaces in the description we still don't count them as a description.
                if request.POST['description'].strip():
                    descriptionValue = request.POST['description']

                # WHEN A USER IS AUTHENTICATED.
                if request.user.is_authenticated:

                    editedIncome = Income.objects.get(id=pk)

                    ## GET ALL THE LANGUAGES FOR THE SOURCE.
                    sourcesLang = Source.objects.get(name=request.POST['source'])
                    
                    editedIncome.name = request.POST['incomeName']
                    editedIncome.date = request.POST['datePicked']
                    editedIncome.description = descriptionValue
                    editedIncome.amount = request.POST['amount']
                    editedIncome.source = sourcesLang.name_en
                    editedIncome.source_en = sourcesLang.name_en
                    editedIncome.source_ja = sourcesLang.name_ja
                    editedIncome.source_es = sourcesLang.name_es

                    editedIncome.save()
                    
                messages.success(request, _("Income edited successfully!"))
                return redirect('income')
        else:
            messages.error(request, _("Naughty Naughty. You don't have permissions to see that."))
            return redirect('income')
    else:
        ## DO SOMETHING WHEN THE USER IS NOT AUTHENTICATED
        if request.method == 'GET':
            return render(request, 'income/edit_income.html', {'sources': sources, 'incomes': pk})
        else:
            messages.success(request, _("Income edited successfully!"))
            return redirect('income')
    
def search_income(request):

    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        search_str = data.get('searchText', '')

        # In case we use redis cache.
        # cache_key = f'search_income_{request.user.id}_{search_str}'
        # cached_result = cache.get(cache_key)

        # if cached_result is not None:
        #     return JsonResponse(cached_result, safe=False)

        income = Income.objects.filter(
            amount__startswith=search_str, user=request.user) | Income.objects.filter(
                date__startswith=search_str, user=request.user) | Income.objects.filter(
                    description__icontains=search_str, user=request.user) | Income.objects.filter(
                        name__icontains=search_str, user=request.user) | Income.objects.filter(
                            source__istartswith=search_str, user=request.user)
        
        preferences = UserPreferences.objects.filter(user=request.user)[0] # Get the currency from the UserPreferences model.
        
        data = list(income.values())
        if data:
            data[0]['currency'] = preferences.currency[:3]
        
        # cache.set(cache_key, data, timeout=300)  # Cache the result for 5 minutes (300 seconds)
        return JsonResponse(list(data), safe=False)
    
    return JsonResponse({'error': _('Invalid request method')}, status=400)

def get_source(income):
    return income.source

def get_source_amount(source, income):
    amount = 0
    bySource = income.filter(source=source)

    for item in bySource:
        amount += item.amount

    return amount

def get_sources(request):
    if request.method == "GET":
        sources = Source.objects.all().values('id', 'name', 'name_en', 'name_es', 'name_ja')
        sources_list = list(sources)  # Convert queryset to list
        return JsonResponse(sources_list, safe=False)

def income_data(request):
    
    if request.method == 'GET':
        # today = datetime.date.today()
        # lastMonth = today.datetime.timedelta(days = 30)
        # last6months = today.datetime.timedelta(days = 30*6)
        # lastyear = today.datetime.timedelta(days = 30*12)
        if request.user.is_authenticated:
            income = Income.objects.filter(user=request.user)

            if income:
                finalRepresentation = {}

                sourceList = list(set(map(get_source, income)))

                for x in income:
                    for y in sourceList:
                        finalRepresentation[y] = get_source_amount(y, income)

                return JsonResponse({'income_data': finalRepresentation}, safe=False)
            else:
                return JsonResponse({'error': _('No income to show.')}, status=404)
        
        else:
            ## DO SOMETHING IF THE USER IS NOT AUTHENTICATED
            return JsonResponse({'error': _('No income to show.')}, status=404)

        return JsonResponse({'error': _('No income to show.')}, status=404)

def income_summary(request):
    if request.method == 'GET':
        if request.user.is_authenticated:
            income = Income.objects.filter(user=request.user)

            if income:
                return render(request, 'income/income_summary.html', {'income': True})
            else:
                return render(request, 'income/income_summary.html')
        
        else:
            ## DO SOMETHING IF THE USER IS NOT AUTHENTICATED
            return render(request, 'income/income_summary.html')
            
    return render(request, 'income/income_summary.html')

def export_data(request):
    if request.user.is_authenticated:
        incomes = Income.objects.filter(user=request.user)
        fileType = request.POST.get('filetype')
        language = translation.get_language()
        timeDate = str(datetime.datetime.now().date())

        match fileType:
            case 'csv':
                response = HttpResponse(content_type='text/csv')
                if language == 'ja':
                    response['Content-Disposition'] = f'attachment; filename = Income-{timeDate}.csv'
                else:
                    response['Content-Disposition'] = _('attachment; filename = Income-{}.csv').format(timeDate)

                writer = csv.writer(response)
                writer.writerow([_('Name'), _('Source'), _('Amount'), _('Description'), _('Date')])

                for income in incomes:
                    writer.writerow([income.name, income.source, income.amount, income.description, income.date])

                return response
            
            case 'xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                if language == 'ja':
                    response['Content-Disposition'] = f'attachment; filename = Income-{timeDate}.xlsx'
                else:
                    response['Content-Disposition'] = _('attachment; filename = Income-{}.xlsx').format(timeDate)

                wb = Workbook()
                ws = wb.active
                ws.title = 'Income'

                # Define the header
                columns = [_('Name'), _('Source'), _('Amount'), _('Description'), _('Date')]

                # Apply the header
                header_font = Font(bold=True)
                for col_num, column_title in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=column_title)
                    cell.font = header_font

                # Write data rows
                rows = incomes.values_list('name', 'source', 'amount', 'description', 'date')
                # print(rows)

                for row_num, row in enumerate(rows, 2):
                    for col_num, cell_value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num, value=str(cell_value))
                
                wb.save(response)
                return response

            
            case 'pdf':
                response = HttpResponse(content_type='text/pdf')
                if language == 'ja':
                    response['Content-Disposition'] = f'attachment; filename = Income-{timeDate}.pdf'
                else:
                    response['Content-Disposition'] = _('attachment; filename = Income-{}.pdf').format(timeDate)
                response['Content-Transfer-Encoding'] = 'binary'

                html_string = render_to_string('income/pdf-output.html', {'incomes': incomes, 'total': incomes.aggregate(Sum('amount'))['amount__sum']})
                html = HTML(string=html_string)

                result = html.write_pdf()

                with tempfile.NamedTemporaryFile(delete=True) as product:
                    product.write(result)
                    product.flush()

                    product= open(product.name, 'rb')
                    response.write(product.read())

                return response

        return HttpResponse(_("Export format not supported"))

    else:
        ## DO SOMETHING IF THE USER IS NOT AUTHENTICATED
        data = json.loads(request.body)
        incomes = data.get('incomes', [])
        fileType = data.get('format')
        # Get Language.
        language = translation.get_language()

        match fileType:
            case 'csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = _('attachment; filename = Income-{}.csv').format(str(datetime.datetime.now().date()))

                writer = csv.writer(response)
                writer.writerow([_('Name'), _('Source'), _('Amount'), _('Description'), _('Date')])

                for income in incomes:
                    if language == 'es':
                        writer.writerow([income['name'], income['source_es'], income['amount'], income['description'], income['date']])
                    elif language == 'ja':
                        writer.writerow([income['name'], income['source_ja'], income['amount'], income['description'], income['date']])
                    else:
                        writer.writerow([income['name'], income['source'], income['amount'], income['description'], income['date']])

                return response
            
            case 'xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = _('attachment; filename = Income-{}.xlsx').format(str(datetime.datetime.now().date()))

                wb = Workbook()
                ws = wb.active
                ws.title = 'Income'

                # Define the header
                columns = [_('Name'), _('Source'), _('Amount'), _('Description'), _('Date')]

                # Apply the header
                header_font = Font(bold=True)
                for col_num, column_title in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=column_title)
                    cell.font = header_font

                # Write data rows
                for row_num, income in enumerate(incomes, 2):
                    ws.cell(row=row_num, column=1, value=income['name'])
                    if language == 'es':
                        ws.cell(row=row_num, column=2, value=income['source_es'])
                    elif language == 'ja':
                        ws.cell(row=row_num, column=2, value=income['source_ja'])
                    else:
                        ws.cell(row=row_num, column=2, value=income['source'])
                    ws.cell(row=row_num, column=3, value=income['amount'])
                    ws.cell(row=row_num, column=4, value=income['description'])
                    ws.cell(row=row_num, column=5, value=income['date'])
                
                wb.save(response)
                return response
 
            case 'pdf':
                response = HttpResponse(content_type='text/pdf')
                response['Content-Disposition'] = _('attachment; filename = Income-{}.pdf').format(str(datetime.datetime.now().date()))
                response['Content-Transfer-Encoding'] = 'binary'

                incomes = data.get('incomes')

                totalAmount = 0.00
                
                for income in incomes:
                    totalAmount += float(income['amount'])
                    if language == 'es':
                        income['source'] = income['source_es']
                    elif language == 'ja':
                        income['source'] = income['source_ja']
                    else:
                        income['source'] = income['source']

                html_string = render_to_string('income/pdf-output.html', {'incomes': incomes, 'total': totalAmount})
                html = HTML(string=html_string)

                result = html.write_pdf()

                with tempfile.NamedTemporaryFile(delete=True) as product:
                    product.write(result)
                    product.flush()

                    product= open(product.name, 'rb')
                    response.write(product.read())

                return response

        return HttpResponse(_("Export format not supported"))
        
    return HttpResponse(_("Export format not supported"))

def import_data(request):
    if request.method == 'POST':
        if request.user.is_authenticated:
            file = request.FILES.get('file')
            delete_previous = request.POST.get('delete_previous') == 'true'
            # print(delete_previous)

            if not file:
                return JsonResponse({'error': _('No file provided')}, status=400)

            file_type = file.name.split('.')[-1].lower()

            try:
                if file_type == 'csv':
                    reader = csv.DictReader(file.read().decode('utf-8').splitlines())
                    df = list(reader)
                elif file_type == 'xls':
                    file_content = file.read()
                    book = xlrd.open_workbook(file_contents=file_content)
                    sheet = book.sheet_by_index(0)
                    headers = sheet.row_values(0)
                    df = [
                        {headers[i]: sheet.cell_value(row, i) for i in range(sheet.ncols)}
                        for row in range(1, sheet.nrows)
                    ]
                elif file_type == 'xlsx':
                    file_content = file.read()
                    wb = load_workbook(BytesIO(file_content))
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    df = [
                        {headers[i]: cell.value for i, cell in enumerate(row)}
                        for row in ws.iter_rows(min_row=2, values_only=False)
                    ]
                else:
                    messages.error(request, _('Unsupported file format'))
                    return JsonResponse({'error': _('Unsupported file format')})
                
                # Check that the file has all the needed columns. From each language.
                if df[0].get('Name') is not None:
                    required_columns = ['Name', 'Source', 'Amount', 'Description', 'Date']
                    lang = 'en'
                
                elif df[0].get('Nombre') is not None:
                    required_columns = ['Nombre', 'Fuente', 'Monto', 'Descripción', 'Fecha']
                    lang = 'es'
                
                elif df[0].get('名前') is not None:
                    required_columns = ['名前', 'ソース', '金額', '説明', '日付']
                    lang = 'ja'
                
                else:
                    messages.error(request, _('Wrong amount of columns or names.'))
                    return JsonResponse({'error': _('Wrong amount of columns or names.')}, status=400)

                if not all(column in df[0] for column in required_columns):
                    messages.error(request, _('Wrong amount of columns or names.'))
                    return JsonResponse({'error': _('Wrong amount of columns or names.')}, status=400)
                
                # Clear existing records if the user requested it.
                if delete_previous:
                    Income.objects.filter(user=request.user).delete()
                
                # We get the language of the page and change it temporarily to the one of the file if needed to be able to import the data.
                currentLang = translation.get_language()
                
                if currentLang != lang:
                    translation.activate(lang)

                # print(df)

                # Process each row in the dataframe.
                for row in df:
                    ## GET ALL THE LANGUAGES FOR THE SOURCE.
                    sourceLang = Source.objects.get(name=row[required_columns[1]])
                    Income.objects.create(
                        user=request.user,
                        name=row[required_columns[0]],
                        source=sourceLang.name_en,
                        source_en =sourceLang.name_en,
                        source_es =sourceLang.name_es,
                        source_ja =sourceLang.name_ja,
                        amount=row[required_columns[2]],
                        description=row[required_columns[3]],
                        date=row[required_columns[4]],
                    )
                
                # If the page language was changed we set it back to what it was.
                if currentLang != lang:
                    translation.activate(currentLang)

                messages.success(request, _('Income imported successfully'))
                return JsonResponse({'success': True})

            except Exception as e:
                messages.error(request, _('Error processing file: {}').format(e))
                return JsonResponse({'error': _('Error processing file: {}').format(e)}, status=400)
        
        else:
            ## DO SOMETHING IF THE USER IS NOT AUTHENTICATED.
            return JsonResponse({'error': _('Error processing file: {}').format(e)}, status=400)
            
    else:
        form = UploadFileForm()